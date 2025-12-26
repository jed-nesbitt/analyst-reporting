from __future__ import annotations

from pathlib import Path
import pandas as pd

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule


HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")  # light blue-grey
HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=12)
H1_FONT = Font(bold=True, size=16)
CENTER = Alignment(horizontal="center")
LEFT = Alignment(horizontal="left")


def _currency_format(currency_code: str) -> str:
    code = (currency_code or "AUD").upper().strip()
    # Works globally: shows currency code as text prefix.
    return f'"{code} " #,##0.00'


PERCENT_FMT = "0.00%"
INT_FMT = "#,##0"
DATE_FMT = "yyyy-mm-dd"
MONTH_FMT = "yyyy-mm"


def _is_percent_col(name: str) -> bool:
    n = name.lower()
    return "margin" in n or n.endswith("_pct")


def _is_date_col(name: str) -> bool:
    n = name.lower()
    return n == "date" or "date" in n


def _is_month_col(name: str) -> bool:
    return name.lower() == "month"


def _is_int_col(name: str) -> bool:
    n = name.lower()
    return n in {"units", "rows_loaded"} or n.endswith("_count")


def _is_currency_col(name: str) -> bool:
    n = name.lower()
    currency_keywords = {"revenue", "sales", "cost", "cogs", "gross_profit", "profit", "amount"}
    if n in currency_keywords:
        return True
    if n.endswith("_mom_abs"):
        return True
    if any(k in n for k in ["revenue", "cost", "gross_profit"]):
        return True
    return False


def _style_header_row(ws, header_row: int, max_col: int) -> None:
    for c in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER


def _auto_fit_columns(ws, min_row: int = 1, max_row: int | None = None, max_col: int | None = None) -> None:
    if max_row is None:
        max_row = ws.max_row
    if max_col is None:
        max_col = ws.max_column

    for col in range(1, max_col + 1):
        max_len = 0
        for row in range(min_row, max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)

        width = min(max_len + 2, 45)
        ws.column_dimensions[get_column_letter(col)].width = max(10, width)


def _apply_formats_by_header(
    ws,
    header_row: int,
    data_start_row: int,
    data_end_row: int,
    currency_fmt: str,
) -> None:
    headers: dict[int, str] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        if v is None:
            continue
        headers[col] = str(v)

    for col, name in headers.items():
        if _is_month_col(name):
            fmt = MONTH_FMT
        elif _is_date_col(name):
            fmt = DATE_FMT
        elif _is_percent_col(name):
            fmt = PERCENT_FMT
        elif _is_int_col(name):
            fmt = INT_FMT
        elif _is_currency_col(name):
            fmt = currency_fmt
        else:
            fmt = None

        if fmt:
            for r in range(data_start_row, data_end_row + 1):
                cell = ws.cell(row=r, column=col)
                if cell.value is None:
                    continue
                if isinstance(cell.value, str) and not (_is_date_col(name) or _is_month_col(name)):
                    continue
                cell.number_format = fmt

        for r in range(data_start_row, data_end_row + 1):
            cell = ws.cell(row=r, column=col)
            if cell.value is None:
                continue
            if isinstance(cell.value, str):
                cell.alignment = LEFT


def _format_summary_sheet(ws, currency_fmt: str) -> None:
    ws.freeze_panes = "A2"
    _style_header_row(ws, header_row=1, max_col=ws.max_column)

    for r in range(2, ws.max_row + 1):
        metric = ws.cell(row=r, column=1).value
        val_cell = ws.cell(row=r, column=2)
        if metric is None or val_cell.value is None:
            continue
        m = str(metric).lower()

        if "margin" in m:
            val_cell.number_format = PERCENT_FMT
        elif m in {"units", "rows_loaded"}:
            val_cell.number_format = INT_FMT
        elif m in {"revenue", "cost", "gross_profit"}:
            val_cell.number_format = currency_fmt
        else:
            if isinstance(val_cell.value, (int, float)):
                val_cell.number_format = "#,##0.00"

    _auto_fit_columns(ws)


def _format_table_sheet(ws, currency_fmt: str) -> None:
    ws.freeze_panes = "A2"
    _style_header_row(ws, header_row=1, max_col=ws.max_column)
    _apply_formats_by_header(
        ws,
        header_row=1,
        data_start_row=2,
        data_end_row=ws.max_row,
        currency_fmt=currency_fmt,
    )
    _auto_fit_columns(ws)


def _write_table(ws, df: pd.DataFrame, title: str, startrow: int, currency_fmt: str) -> int:
    title_row = startrow + 1
    header_row = startrow + 3
    data_start = startrow + 4

    ws.cell(row=title_row, column=1, value=title).font = TITLE_FONT

    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=header_row, column=j, value=str(col))

    for i in range(len(df)):
        for j, col in enumerate(df.columns, start=1):
            val = df.iloc[i, j - 1]
            ws.cell(row=data_start + i, column=j, value=None if pd.isna(val) else val)

    _style_header_row(ws, header_row=header_row, max_col=len(df.columns))
    if len(df) > 0:
        _apply_formats_by_header(
            ws,
            header_row=header_row,
            data_start_row=data_start,
            data_end_row=data_start + len(df) - 1,
            currency_fmt=currency_fmt,
        )

    return startrow + len(df) + 6


def _safe_latest_month(trends: pd.DataFrame) -> pd.Timestamp | None:
    if "month" not in trends.columns or trends.empty:
        return None
    s = pd.to_datetime(trends["month"], errors="coerce").dropna()
    if s.empty:
        return None
    return s.max()


def _build_exec_summary_rows(
    summary: pd.DataFrame,
    trends: pd.DataFrame,
    variance: pd.DataFrame,
    drill_region: pd.DataFrame,
    drill_product: pd.DataFrame,
    currency_code: str,
) -> list[list[object]]:
    cur = (currency_code or "AUD").upper().strip()

    m = dict(zip(summary["metric"], summary["value"]))
    rows_loaded = m.get("rows_loaded", pd.NA)

    latest = _safe_latest_month(trends)

    # MoM from variance (use last non-null row)
    last_var = variance.copy()
    if "month" in last_var.columns:
        last_var["month"] = pd.to_datetime(last_var["month"], errors="coerce")
        last_var = last_var.sort_values("month")

    def _last_value(col: str):
        if col not in last_var.columns:
            return pd.NA
        s = pd.to_numeric(last_var[col], errors="coerce").dropna()
        return s.iloc[-1] if len(s) else pd.NA

    rev_mom_pct = _last_value("revenue_mom_pct")
    margin_mom_abs = _last_value("margin_mom_abs")

    # Latest-month top region/product by revenue
    top_region = ""
    top_region_rev = pd.NA
    if latest is not None and not drill_region.empty and {"month", "region", "revenue"}.issubset(drill_region.columns):
        dr = drill_region.copy()
        dr["month"] = pd.to_datetime(dr["month"], errors="coerce")
        dr = dr[dr["month"] == latest].copy()
        if not dr.empty:
            dr["revenue"] = pd.to_numeric(dr["revenue"], errors="coerce")
            dr = dr.dropna(subset=["revenue"])
            if not dr.empty:
                r = dr.loc[dr["revenue"].idxmax()]
                top_region = str(r.get("region", "Unknown"))
                top_region_rev = r["revenue"]

    top_product = ""
    top_product_rev = pd.NA
    if latest is not None and not drill_product.empty and {"month", "product", "revenue"}.issubset(drill_product.columns):
        dp = drill_product.copy()
        dp["month"] = pd.to_datetime(dp["month"], errors="coerce")
        dp = dp[dp["month"] == latest].copy()
        if not dp.empty:
            dp["revenue"] = pd.to_numeric(dp["revenue"], errors="coerce")
            dp = dp.dropna(subset=["revenue"])
            if not dp.empty:
                r = dp.loc[dp["revenue"].idxmax()]
                top_product = str(r.get("product", "Unknown"))
                top_product_rev = r["revenue"]

    latest_label = latest.strftime("%Y-%m") if latest is not None else "N/A"

    rows: list[list[object]] = []
    rows.append(["Insight", "Value"])
    rows.append(["Latest month", latest_label])
    rows.append(["Rows loaded", rows_loaded])

    # Revenue MoM %
    if pd.notna(rev_mom_pct):
        rows.append(["Revenue MoM", float(rev_mom_pct)])  # format later as percent
    else:
        rows.append(["Revenue MoM", "N/A"])

    # Margin change (absolute)
    if pd.notna(margin_mom_abs):
        # margin_mom_abs is in ratio terms; show in bps
        rows.append(["Margin change (MoM)", float(margin_mom_abs)])  # format later as percent or bps-like
    else:
        rows.append(["Margin change (MoM)", "N/A"])

    if top_region:
        rows.append([f"Top region by revenue ({latest_label})", top_region])
        rows.append([f"Top region revenue ({latest_label})", top_region_rev])
    else:
        rows.append([f"Top region by revenue ({latest_label})", "N/A"])

    if top_product:
        rows.append([f"Top product by revenue ({latest_label})", top_product])
        rows.append([f"Top product revenue ({latest_label})", top_product_rev])
    else:
        rows.append([f"Top product by revenue ({latest_label})", "N/A"])

    rows.append(["Currency", cur])
    return rows


def _write_exec_summary_sheet(
    book,
    *,
    summary: pd.DataFrame,
    trends: pd.DataFrame,
    variance: pd.DataFrame,
    drill_region: pd.DataFrame,
    drill_product: pd.DataFrame,
    currency_code: str,
    currency_fmt: str,
) -> None:
    # Insert at the front
    if "ExecutiveSummary" in book.sheetnames:
        ws = book["ExecutiveSummary"]
        book.remove(ws)

    ws = book.create_sheet("ExecutiveSummary", 0)

    ws["A1"] = "Executive Summary"
    ws["A1"].font = H1_FONT

    ws["A2"] = "Auto-generated highlights for stakeholders"
    ws["A2"].alignment = LEFT

    rows = _build_exec_summary_rows(summary, trends, variance, drill_region, drill_product, currency_code)

    start_row = 4
    for i, (insight, value) in enumerate(rows):
        r = start_row + i
        ws.cell(row=r, column=1, value=insight)
        ws.cell(row=r, column=2, value=value)

    # Header style
    _style_header_row(ws, header_row=start_row, max_col=2)

    # Formats for specific rows
    # Revenue MoM row: percent
    for r in range(start_row + 1, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        val_cell = ws.cell(row=r, column=2)

        if label == "Revenue MoM" and isinstance(val_cell.value, (int, float)):
            val_cell.number_format = PERCENT_FMT

        if label == "Margin change (MoM)" and isinstance(val_cell.value, (int, float)):
            # show margin change as bps-like percent (e.g., -0.0045 -> -0.45%)
            val_cell.number_format = PERCENT_FMT

        if isinstance(label, str) and "revenue (" in label.lower() and isinstance(val_cell.value, (int, float)):
            val_cell.number_format = currency_fmt

        if label == "Rows loaded" and isinstance(val_cell.value, (int, float)):
            val_cell.number_format = INT_FMT

    ws.freeze_panes = f"A{start_row+1}"
    _auto_fit_columns(ws)


def _apply_variance_conditional_formatting(ws) -> None:
    """
    Apply red↔green color scale to *_mom_abs and *_mom_pct columns (centered at 0).
    """
    if ws.max_row < 3:
        return

    max_row = ws.max_row
    max_col = ws.max_column

    # Excel-style red/yellow/green scale
    rule = ColorScaleRule(
        start_type="min",
        start_color="F8696B",  # red
        mid_type="num",
        mid_value=0,
        mid_color="FFEB84",    # yellow
        end_type="max",
        end_color="63BE7B",    # green
    )

    for col in range(1, max_col + 1):
        header = ws.cell(row=1, column=col).value
        if header is None:
            continue
        h = str(header).lower()
        if h.endswith("_mom_abs") or h.endswith("_mom_pct"):
            col_letter = get_column_letter(col)
            rng = f"{col_letter}2:{col_letter}{max_row}"
            ws.conditional_formatting.add(rng, rule)


def write_excel_pack(
    out_path: Path,
    summary: pd.DataFrame,
    trends: pd.DataFrame,
    variance: pd.DataFrame,
    drill_region: pd.DataFrame,
    drill_product: pd.DataFrame,
    warnings: list[str] | None = None,
    currency_code: str = "AUD",
) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    currency_fmt = _currency_format(currency_code)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # Write main tables
        summary.to_excel(writer, sheet_name="Summary", index=False)
        trends.to_excel(writer, sheet_name="Trends", index=False)
        variance.to_excel(writer, sheet_name="Variance", index=False)

        book = writer.book

        # Executive summary tab (uses the same dataframes you already computed)
        _write_exec_summary_sheet(
            book,
            summary=summary,
            trends=trends,
            variance=variance,
            drill_region=drill_region,
            drill_product=drill_product,
            currency_code=currency_code,
            currency_fmt=currency_fmt,
        )

        # Format the pandas-written sheets
        ws_sum = writer.sheets["Summary"]
        ws_tr = writer.sheets["Trends"]
        ws_var = writer.sheets["Variance"]

        _format_summary_sheet(ws_sum, currency_fmt=currency_fmt)
        _format_table_sheet(ws_tr, currency_fmt=currency_fmt)
        _format_table_sheet(ws_var, currency_fmt=currency_fmt)

        # Conditional formatting for variance
        _apply_variance_conditional_formatting(ws_var)

        # Drilldowns tab
        ws = book.create_sheet("Drilldowns")

        r = 0
        if warnings:
            ws.cell(row=1, column=1, value="WARNINGS").font = TITLE_FONT
            for i, w in enumerate(warnings, start=2):
                ws.cell(row=i, column=1, value=w)
            r = len(warnings) + 3

        r = _write_table(ws, drill_region, "Revenue & GP by Month x Region", startrow=r, currency_fmt=currency_fmt)
        r = _write_table(ws, drill_product, "Revenue & GP by Month x Product", startrow=r, currency_fmt=currency_fmt)

        _auto_fit_columns(ws)

        if "Sheet" in book.sheetnames:
            book.remove(book["Sheet"])
